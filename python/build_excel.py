"""
Project 2: Personal Financial Intelligence System
Step 5 - Excel dashboard (openpyxl)

Builds excel/Financial_Dashboard.xlsx with:
  - Raw Data sheet (all cleaned transactions - ready for native Excel PivotTables)
  - Dashboard sheet: live formulas (SUMIFS, AVERAGEIFS, XLOOKUP, IF) answering
    every dashboard question, plus charts
  - Instructions sheet

Newer Excel functions (XLOOKUP) are written with the required "_xlfn." prefix -
without it, formulas written by a non-Excel tool show #NAME? in real Excel even
though the function itself is valid. Learned and fixed during Project 1's build.
"""

import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.chart import BarChart, LineChart, Reference
from openpyxl.formatting.rule import ColorScaleRule
from openpyxl.worksheet.table import Table, TableStyleInfo

CLEAN = "/home/claude/project2/data/clean"
OUT = "/home/claude/project2/excel/Financial_Dashboard.xlsx"

HEADER_FILL = PatternFill("solid", fgColor="1F2937")
HEADER_FONT = Font(color="FFFFFF", bold=True)
TITLE_FONT = Font(size=16, bold=True, color="1F2937")
KPI_LABEL_FONT = Font(size=10, color="6B7280")
KPI_VALUE_FONT = Font(size=20, bold=True, color="2563EB")
KPI_VALUE_WARN_FONT = Font(size=20, bold=True, color="DC2626")
SECTION_FONT = Font(size=12, bold=True, color="1F2937")


def style_header_row(ws, row, ncols):
    for c in range(1, ncols + 1):
        cell = ws.cell(row=row, column=c)
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = Alignment(horizontal="center", vertical="center")


def autosize(ws, df):
    for i, col in enumerate(df.columns, start=1):
        max_len = 10
        if len(df):
            max_len = max(len(str(v)) for v in df[col].tolist())
        width = max(12, min(38, max_len + 2, len(str(col)) + 4))
        ws.column_dimensions[get_column_letter(i)].width = width


def write_df_as_table(ws, df, table_name, start_row=1):
    for j, col in enumerate(df.columns, start=1):
        ws.cell(row=start_row, column=j, value=col)
    for i, row in enumerate(df.itertuples(index=False), start=start_row + 1):
        for j, val in enumerate(row, start=1):
            ws.cell(row=i, column=j, value=val)
    style_header_row(ws, start_row, len(df.columns))
    autosize(ws, df)
    last_row = start_row + len(df)
    last_col = get_column_letter(len(df.columns))
    ref = f"A{start_row}:{last_col}{last_row}"
    table = Table(displayName=table_name, ref=ref)
    table.tableStyleInfo = TableStyleInfo(name="TableStyleMedium2", showRowStripes=True)
    ws.add_table(table)
    return last_row


def main():
    txns = pd.read_csv(f"{CLEAN}/transactions.csv")
    raw = txns[["transaction_id", "date", "account", "merchant", "category",
                "description_raw", "amount", "txn_type", "month", "day_of_week"]].copy()
    raw.columns = ["Transaction ID", "Date", "Account", "Merchant", "Category",
                   "Description (raw)", "Amount", "Type", "Month", "Day of Week"]

    wb = Workbook()
    ws_raw = wb.active
    ws_raw.title = "Raw Data"
    last_row = write_df_as_table(ws_raw, raw, "TransactionsTable")
    n = last_row
    RD = "'Raw Data'"

    ws = wb.create_sheet("Dashboard", 0)
    ws.sheet_view.showGridLines = False
    ws.column_dimensions["A"].width = 2
    for col in "BCDEFGHI":
        ws.column_dimensions[col].width = 15

    ws["B2"] = "Personal Financial Intelligence Dashboard"
    ws["B2"].font = TITLE_FONT
    ws["B3"] = "All figures below are live Excel formulas (SUMIFS / AVERAGEIFS / XLOOKUP / IF) referencing the Raw Data sheet."
    ws["B3"].font = Font(italic=True, size=9, color="6B7280")

    # --- KPI row ---
    kpis = [
        ("Total Income", f'=SUMIFS({RD}!G2:G{n},{RD}!G2:G{n},">0")', False),
        ("Total Expenses", f'=-SUMIFS({RD}!G2:G{n},{RD}!G2:G{n},"<0")', False),
        ("Net (Income - Expenses)",
         f'=SUMIFS({RD}!G2:G{n},{RD}!G2:G{n},">0")+SUMIFS({RD}!G2:G{n},{RD}!G2:G{n},"<0")', True),
        ("Savings Rate",
         f'=TEXT((SUMIFS({RD}!G2:G{n},{RD}!G2:G{n},">0")+SUMIFS({RD}!G2:G{n},{RD}!G2:G{n},"<0"))'
         f'/SUMIFS({RD}!G2:G{n},{RD}!G2:G{n},">0"),"0.0%")', True),
    ]
    for i, (label, formula, is_net) in enumerate(kpis):
        col = 2 + i * 2
        ws.cell(row=5, column=col, value=label).font = KPI_LABEL_FONT
        cell = ws.cell(row=6, column=col, value=formula)
        cell.font = KPI_VALUE_FONT
        if col in (2, 4):
            cell.number_format = '"$"#,##0'

    # --- Spending by category (Where does most of my money go?) ---
    ws["B9"] = "Spending by Category (Where does most of my money go?)"
    ws["B9"].font = SECTION_FONT
    categories = sorted(raw["Category"].unique())
    ws.cell(row=10, column=2, value="Category")
    ws.cell(row=10, column=3, value="Total Spent")
    ws.cell(row=10, column=4, value="% of Spending")
    style_header_row(ws, 10, 3)
    for i, cat in enumerate(categories, start=11):
        ws.cell(row=i, column=2, value=cat)
        ws.cell(row=i, column=3,
                value=f'=-SUMIFS({RD}!G2:G{n},{RD}!E2:E{n},B{i},{RD}!G2:G{n},"<0")')
        ws.cell(row=i, column=3).number_format = '"$"#,##0'
        ws.cell(row=i, column=4, value=f'=C{i}/SUMIFS({RD}!$G$2:$G${n},{RD}!$G$2:$G${n},"<0")*-1')
        ws.cell(row=i, column=4).number_format = "0.0%"
    cat_last_row = 10 + len(categories)

    ws.cell(row=cat_last_row + 2, column=2, value="Biggest category (XLOOKUP by highest spend):")
    ws.cell(row=cat_last_row + 2, column=2).font = Font(italic=True, size=9)
    ws.cell(row=cat_last_row + 3, column=2,
            value=f"=_xlfn.XLOOKUP(MAX(C11:C{cat_last_row}),C11:C{cat_last_row},B11:B{cat_last_row})")
    ws.cell(row=cat_last_row + 3, column=2).font = Font(bold=True, color="2563EB")
    ws.cell(row=cat_last_row + 4, column=2,
            value=(f"(Excel 2019/365 or Google Sheets only. Older Excel: "
                    f"=INDEX(B11:B{cat_last_row},MATCH(MAX(C11:C{cat_last_row}),C11:C{cat_last_row},0)) )"))
    ws.cell(row=cat_last_row + 4, column=2).font = Font(italic=True, size=8, color="9CA3AF")

    # --- Top merchants (Which merchants receive the most money?) ---
    merch_start = cat_last_row + 7
    ws.cell(row=merch_start - 1, column=2, value="Top Merchants (Which merchants receive the most money?)")
    ws.cell(row=merch_start - 1, column=2).font = SECTION_FONT
    top_merchants = (txns[txns["amount"] < 0].groupby("merchant")["amount"].sum().abs()
                      .sort_values(ascending=False).head(10))
    ws.cell(row=merch_start, column=2, value="Merchant")
    ws.cell(row=merch_start, column=3, value="Total Spent")
    ws.cell(row=merch_start, column=4, value="Visits")
    style_header_row(ws, merch_start, 3)
    for i, (merch, _amt) in enumerate(top_merchants.items(), start=merch_start + 1):
        ws.cell(row=i, column=2, value=merch)
        ws.cell(row=i, column=3, value=f'=-SUMIFS({RD}!G2:G{n},{RD}!D2:D{n},B{i},{RD}!G2:G{n},"<0")')
        ws.cell(row=i, column=3).number_format = '"$"#,##0'
        ws.cell(row=i, column=4, value=f'=COUNTIFS({RD}!D2:D{n},B{i},{RD}!G2:G{n},"<0")')
    merch_last_row = merch_start + len(top_merchants)

    # --- Subscriptions (What subscriptions am I paying for?) ---
    sub_start = merch_last_row + 3
    ws.cell(row=sub_start - 1, column=2, value="Active Subscriptions (What subscriptions am I paying for?)")
    ws.cell(row=sub_start - 1, column=2).font = SECTION_FONT
    subs = sorted(txns[txns["category"] == "Subscriptions"]["merchant"].unique())
    ws.cell(row=sub_start, column=2, value="Subscription")
    ws.cell(row=sub_start, column=3, value="Monthly Cost")
    ws.cell(row=sub_start, column=4, value="Est. Annual Cost")
    style_header_row(ws, sub_start, 3)
    for i, sub in enumerate(subs, start=sub_start + 1):
        ws.cell(row=i, column=2, value=sub)
        ws.cell(row=i, column=3, value=f'=-AVERAGEIFS({RD}!G2:G{n},{RD}!D2:D{n},B{i})')
        ws.cell(row=i, column=3).number_format = '"$"#,##0.00'
        ws.cell(row=i, column=4, value=f'=C{i}*12')
        ws.cell(row=i, column=4).number_format = '"$"#,##0'
    sub_last_row = sub_start + len(subs)
    ws.cell(row=sub_last_row + 1, column=2, value="Total")
    ws.cell(row=sub_last_row + 1, column=2).font = Font(bold=True)
    ws.cell(row=sub_last_row + 1, column=3, value=f"=SUM(C{sub_start+1}:C{sub_last_row})")
    ws.cell(row=sub_last_row + 1, column=3).number_format = '"$"#,##0.00'
    ws.cell(row=sub_last_row + 1, column=3).font = Font(bold=True)
    ws.cell(row=sub_last_row + 1, column=4, value=f"=SUM(D{sub_start+1}:D{sub_last_row})")
    ws.cell(row=sub_last_row + 1, column=4).number_format = '"$"#,##0'
    ws.cell(row=sub_last_row + 1, column=4).font = Font(bold=True)

    # --- Monthly income vs expense (trend + net) ---
    month_start = sub_last_row + 4
    ws.cell(row=month_start - 1, column=2, value="Monthly Income vs. Expenses")
    ws.cell(row=month_start - 1, column=2).font = SECTION_FONT
    months = sorted(raw["Month"].unique())
    ws.cell(row=month_start, column=2, value="Month")
    ws.cell(row=month_start, column=3, value="Income")
    ws.cell(row=month_start, column=4, value="Expenses")
    ws.cell(row=month_start, column=5, value="Net")
    style_header_row(ws, month_start, 4)
    for i, m in enumerate(months, start=month_start + 1):
        ws.cell(row=i, column=2, value=m)
        ws.cell(row=i, column=3,
                value=f'=SUMIFS({RD}!G2:G{n},{RD}!I2:I{n},B{i},{RD}!G2:G{n},">0")')
        ws.cell(row=i, column=3).number_format = '"$"#,##0'
        ws.cell(row=i, column=4,
                value=f'=-SUMIFS({RD}!G2:G{n},{RD}!I2:I{n},B{i},{RD}!G2:G{n},"<0")')
        ws.cell(row=i, column=4).number_format = '"$"#,##0'
        ws.cell(row=i, column=5, value=f'=C{i}-D{i}')
        ws.cell(row=i, column=5).number_format = '"$"#,##0'
    month_last_row = month_start + len(months)

    ws.conditional_formatting.add(
        f"C11:C{cat_last_row}",
        ColorScaleRule(start_type="min", start_color="86EFAC", end_type="max", end_color="FCA5A5"))
    ws.conditional_formatting.add(
        f"E{month_start+1}:E{month_last_row}",
        ColorScaleRule(start_type="min", start_color="FCA5A5", end_type="max", end_color="86EFAC"))

    # --- charts ---
    bar1 = BarChart()
    bar1.title = "Spending by Category"
    data = Reference(ws, min_col=3, min_row=10, max_row=cat_last_row)
    cats = Reference(ws, min_col=2, min_row=11, max_row=cat_last_row)
    bar1.add_data(data, titles_from_data=True)
    bar1.set_categories(cats)
    bar1.height, bar1.width = 9, 15
    ws.add_chart(bar1, "G5")

    bar2 = BarChart()
    bar2.title = "Top Merchants by Spend"
    data2 = Reference(ws, min_col=3, min_row=merch_start, max_row=merch_last_row)
    cats2 = Reference(ws, min_col=2, min_row=merch_start + 1, max_row=merch_last_row)
    bar2.add_data(data2, titles_from_data=True)
    bar2.set_categories(cats2)
    bar2.height, bar2.width = 9, 15
    ws.add_chart(bar2, "G24")

    line1 = LineChart()
    line1.title = "Monthly Income vs. Expenses"
    data3 = Reference(ws, min_col=3, max_col=4, min_row=month_start, max_row=month_last_row)
    cats3 = Reference(ws, min_col=2, min_row=month_start + 1, max_row=month_last_row)
    line1.add_data(data3, titles_from_data=True)
    line1.set_categories(cats3)
    line1.height, line1.width = 9, 15
    ws.add_chart(line1, "G43")

    # ---------------- Instructions sheet ----------------
    ws_info = wb.create_sheet("Instructions")
    ws_info.column_dimensions["B"].width = 100
    lines = [
        "PERSONAL FINANCIAL INTELLIGENCE SYSTEM - EXCEL DASHBOARD",
        "",
        "Sheets:",
        "  Dashboard    Live KPIs, category/merchant/subscription breakdowns, and charts.",
        "  Raw Data     Every transaction, as an Excel Table (TransactionsTable) - add new rows here.",
        "",
        "How the formulas work:",
        "  - SUMIFS / AVERAGEIFS / COUNTIFS drive every KPI and breakout table.",
        "  - Amounts are signed: negative = money out, positive = money in - so \"<0\"/\">0\"",
        "    criteria split spending from income throughout.",
        "  - XLOOKUP finds the single biggest spending category automatically",
        "    (requires Excel 2019/365 or Google Sheets - an INDEX/MATCH fallback is noted next to it).",
        "",
        "To build a native Excel PivotTable (optional, in addition to the formula dashboard above):",
        "  1. Click any cell inside the Raw Data table.",
        "  2. Insert > PivotTable > New Worksheet.",
        "  3. Drag 'Category' to Rows and 'Amount' to Values (Sum).",
        "  4. Repeat for Merchant, Month, or Account to explore other cuts.",
        "",
        "This workbook was generated as part of a 3-project data analyst portfolio",
        "(Job Application Intelligence -> Personal Financial Intelligence -> Apartment Decision Engine).",
        "The anomaly detection, next-month forecast, financial health score, and savings",
        "recommendations (the project's advanced features) run in Python - see python/analysis.py",
        "and the README for the full output.",
    ]
    for i, line in enumerate(lines, start=2):
        cell = ws_info.cell(row=i, column=2, value=line)
        if i == 2:
            cell.font = TITLE_FONT
        elif line.strip().endswith(":") and not line.startswith("  "):
            cell.font = SECTION_FONT

    wb.save(OUT)
    print(f"Saved {OUT}")


if __name__ == "__main__":
    main()
